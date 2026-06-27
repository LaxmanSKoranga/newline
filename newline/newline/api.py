import io
import re

import frappe
from frappe.utils import flt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter


def _strip_html(s):
	if not s:
		return ""
	s = str(s)
	s = re.sub(r'<br\s*/?>', '\n', s, flags=re.IGNORECASE)
	s = re.sub(r'</p\s*>', '\n', s, flags=re.IGNORECASE)
	s = re.sub(r'</div\s*>', '\n', s, flags=re.IGNORECASE)
	s = re.sub(r'<[^>]+>', '', s)
	s = (s.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
		  .replace('&nbsp;', ' ').replace('&#39;', "'").replace('&quot;', '"'))
	s = re.sub(r'\n{3,}', '\n\n', s)
	return s.strip()

_COLS = [
	("A",  "nl_is",                   "IS",            36,  False),
	("B",  "nl_product_package",      "PKG",           80,  False),
	("C",  "nl_specification",        "SPEC",          95,  False),
	("D",  "item_code",                "TYPE",          85,  False),
	("E",  "nl_location",             "LOCATION",      80,  False),
	("F",  "nl_proposed_brand",       "BRAND",         80,  False),
	("G",  "nl_proposed_product",     "PRODUCT",      105,  False),
	("H",  "description",             "DESCRIPTION",  190,  False),
	("I",  "nl_price_type",           "PRICE TYPE",    68,  False),
	("J",  "nl_supplier_brand",       "SUP.BRAND",     80,  False),
	("K",  "nl_uexw_value",           "U.EXW",         78,  False),
	("L",  "nl_discount_pct",         "DISC%",         50,  False),
	("M",  "nl_net_uexw",             "NET EXW",       78,  True),
	("N",  "nl_exw_currency",         "CUR",           36,  False),
	("O",  "nl_fx_rate",              "FX",            58,  False),
	("P",  "nl_exworks_unit_aed",     "UNIT AED",      82,  True),
	("Q",  "nl_exworks_total_aed",    "TOT AED",       82,  True),
	("R",  "nl_ship_pct",             "SHIP%",         40,  False),
	("S",  "nl_ship_unit_aed",        "UNIT",          72,  True),
	("T",  "nl_ship_total_aed",       "TOTAL",         72,  True),
	("U",  "nl_ins_pct",              "INS%",          40,  False),
	("V",  "nl_ins_unit_aed",         "UNIT",          72,  True),
	("W",  "nl_ins_total_aed",        "TOTAL",         72,  True),
	("X",  "nl_cus_pct",              "CUS%",          40,  False),
	("Y",  "nl_cus_unit_aed",         "UNIT",          72,  True),
	("Z",  "nl_cus_total_aed",        "TOTAL",         72,  True),
	("AA", "nl_sam_pct",              "SAM%",          40,  False),
	("AB", "nl_sam_unit_aed",         "UNIT",          72,  True),
	("AC", "nl_sam_total_aed",        "TOTAL",         72,  True),
	("AD", "nl_lc_pct",               "LC%",           40,  False),
	("AE", "nl_lc_unit_aed",          "UNIT",          72,  True),
	("AF", "nl_lc_total_aed",         "TOTAL",         72,  True),
	("AG", "nl_landed_unit_aed",      "UNIT AED",      84,  True),
	("AH", "nl_landed_total_aed",     "TOT AED",       84,  True),
	("AI", "nl_markup",               "MU",            42,  False),
	("AJ", "nl_unit_sell_aed",        "UNIT SELL",     84,  True),
	("AK", "nl_total_sell_aed",       "TOT SELL",      84,  True),
	("AL", "nl_gm_pct",               "GM%",           50,  True),
	("AM", "qty",                      "QTY",           48,  False),
	("AN", "uom",                      "UOM",           42,  False),
	("AO", "nl_approval_risk",        "RISK",          58,  False),
	("AP", "nl_alt1_brand",           "ALT1 BRAND",    75,  False),
	("AQ", "nl_alt1_product",         "ALT1 PRODUCT",  92,  False),
	("AR", "nl_alt2_brand",           "ALT2 BRAND",    75,  False),
	("AS", "nl_alt2_product",         "ALT2 PRODUCT",  92,  False),
]

_GROUPS = [
	("A",  "J",  "PROPOSED PRODUCT DETAILS", "C5CAE9", "1A1A3E"),
	("K",  "Q",  "EXWORKS",                  "BBDEFB", "1E3A5F"),
	("R",  "T",  "FREIGHT",                  "C8E6C9", "1A4731"),
	("U",  "W",  "INSURANCE",                "FFF9C4", "3D3416"),
	("X",  "Z",  "CUSTOMS",                  "FFCDD2", "4A1E1E"),
	("AA", "AC", "SAMPLES",                  "E1BEE7", "2D1A4A"),
	("AD", "AF", "LETTER OF CREDIT",         "B2EBF2", "1A3A3A"),
	("AG", "AH", "LANDED AED",               "BBDEFB", "0A2A4A"),
	("AI", "AL", "SELLING",                  "C8E6C9", "1A3D1A"),
	("AM", "AN", "QTY",                      "FFE0B2", "3D2000"),
	("AO", "AO", "RISK",                     "FCE4EC", "3D1A2A"),
	("AP", "AS", "SPECIFICATIONS",           "FBE9E7", "5C1A00"),
]

_LAST_COL  = "AS"
_LANDED    = {"AG", "AH"}
_SELL      = {"AJ", "AK"}
_FX        = {"O"}
_ROW_BG    = {"Main Item": "FFFFFF", "Driver": "EEF4FF", "Accessory": "F8F8F8"}
_PCT_COLS  = {"nl_discount_pct","nl_ship_pct","nl_ins_pct","nl_cus_pct","nl_sam_pct","nl_lc_pct"}
_CCY_COLS  = {
	"nl_uexw_value","nl_net_uexw","nl_tot_exw",
	"nl_exworks_unit_aed","nl_exworks_total_aed",
	"nl_ship_unit_aed","nl_ship_total_aed",
	"nl_ins_unit_aed","nl_ins_total_aed",
	"nl_cus_unit_aed","nl_cus_total_aed",
	"nl_sam_unit_aed","nl_sam_total_aed",
	"nl_lc_unit_aed","nl_lc_total_aed",
	"nl_landed_unit_aed","nl_landed_total_aed",
	"nl_unit_sell_aed","nl_total_sell_aed",
}

def _fill(hex6):
	return PatternFill(fgColor=hex6.lstrip("#"), fill_type="solid")

def _font(bold=False, color="000000", size=9):
	return Font(bold=bold, color=color.lstrip("#"), size=size)

def _thin():
	s = Side(style="thin", color="D0D0D0")
	return Border(left=s, right=s, top=s, bottom=s)

def _left_accent(color_hex):
	s = Side(style="thin", color="D0D0D0")
	return Border(
		left=Side(style="thick", color=color_hex),
		right=s, top=s, bottom=s,
	)

def _align(h="left", wrap=False):
	return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _formula(field, r):
	return {
		"nl_net_uexw":          f"=K{r}*(1-L{r}/100)",
		"nl_exworks_unit_aed":  f"=M{r}*O{r}",
		"nl_exworks_total_aed": f"=P{r}*AM{r}",
		"nl_ship_unit_aed":     f"=P{r}*R{r}/100",
		"nl_ship_total_aed":    f"=S{r}*AM{r}",
		"nl_ins_unit_aed":      f"=P{r}*U{r}/100",
		"nl_ins_total_aed":     f"=V{r}*AM{r}",
		"nl_cus_unit_aed":      f"=P{r}*X{r}/100",
		"nl_cus_total_aed":     f"=Y{r}*AM{r}",
		"nl_sam_unit_aed":      f"=P{r}*AA{r}/100",
		"nl_sam_total_aed":     f"=AB{r}*AM{r}",
		"nl_lc_unit_aed":       f"=P{r}*AD{r}/100",
		"nl_lc_total_aed":      f"=AE{r}*AM{r}",
		"nl_landed_unit_aed":   f"=P{r}+S{r}+V{r}+Y{r}+AB{r}+AE{r}",
		"nl_landed_total_aed":  f"=AG{r}*AM{r}",
		"nl_unit_sell_aed":     f"=ROUND(AG{r}*AI{r},0)",
		"nl_total_sell_aed":    f"=AJ{r}*AM{r}",
		"nl_gm_pct":            f"=IF(AJ{r}>0,(AJ{r}-AG{r})/AJ{r}*100,0)",
	}.get(field, "")


def _num_fmt(field):
	if field in _CCY_COLS:      return "#,##0.00"
	if field in _PCT_COLS:      return "0.0"
	if field == "nl_gm_pct":   return "0.0"
	if field == "nl_fx_rate":  return "0.0000"
	if field == "nl_markup":   return "0.00"
	if field == "qty":         return "#,##0.##"
	return "@"


def _build_workbook(doc, hidden_groups=None):
	hidden_groups = set(hidden_groups or [])
	wb = Workbook()
	ws = wb.active
	ws.title = "Costing Sheet"

	ws.merge_cells(f"A1:{_LAST_COL}1")
	c = ws["A1"]
	c.value     = "NEWLINE MEA  —  INTERNAL COSTING SHEET"
	c.fill      = _fill("1A1A3E")
	c.font      = Font(bold=True, color="FFFFFF", size=14)
	c.alignment = _align("center")
	ws.row_dimensions[1].height = 28

	proj = doc.nl_project_name  or ""
	ref  = doc.nl_ref_number    or doc.name
	date = str(doc.nl_project_date or doc.transaction_date or "")
	cust = doc.customer_name    or ""
	ws.merge_cells("A2:J2")
	c = ws["A2"]
	c.value     = f"PROJECT: {proj}   |   REF: {ref}   |   CUSTOMER: {cust}   |   DATE: {date}"
	c.font      = Font(bold=True, size=9, color="1A1A3E")
	c.alignment = _align("left")
	c.fill      = _fill("E8EAF0")
	ws["K2"].value = "EXCHANGE RATES:"
	ws["K2"].font  = Font(bold=True, size=8)
	for i, xr in enumerate(doc.nl_exchange_rates or []):
		lc = get_column_letter(column_index_from_string("L") + i * 2)
		lv = get_column_letter(column_index_from_string("L") + i * 2 + 1)
		ws[f"{lc}2"].value = xr.currency
		ws[f"{lc}2"].font  = Font(bold=True, size=8)
		ws[f"{lv}2"].value = flt(xr.rate)
		ws[f"{lv}2"].number_format = "0.0000"
		ws[f"{lv}2"].font  = Font(size=8)
	ws.row_dimensions[2].height = 18

	for c1, c2, label, bg_color, fg_color in _GROUPS:
		ws.merge_cells(f"{c1}3:{c2}3") if c1 != c2 else None
		cell = ws[f"{c1}3"]
		cell.value     = label
		cell.fill      = _fill(bg_color)
		cell.font      = Font(bold=True, color=fg_color, size=8)
		cell.alignment = _align("center")
	ws.row_dimensions[3].height = 20

	col_max = {}
	for col_l, field, label, width_px, _ in _COLS:
		cell = ws[f"{col_l}4"]
		cell.value     = label
		cell.fill      = _fill("2C3E50")
		cell.font      = Font(bold=True, color="ECF0F1", size=8)
		cell.alignment = _align("center")
		col_max[col_l] = len(label) + 2
	ws.row_dimensions[4].height = 18

	ws.freeze_panes = "A5"

	DATA_START = 5
	lines = doc.items or []

	# Direct SQL — get descriptions exactly as stored in DB (bypasses Frappe doc caching)
	_row_desc = {}
	for r_ in frappe.db.sql(
		"SELECT name, description FROM `tabQuotation Item` WHERE parent=%s",
		doc.name, as_dict=True
	):
		_row_desc[r_["name"]] = _strip_html(r_.get("description") or "")

	# Fallback: item master descriptions
	_item_codes = list({l.item_code for l in lines if l.item_code})
	_item_desc  = {}
	if _item_codes:
		placeholders = ", ".join(["%s"] * len(_item_codes))
		for r_ in frappe.db.sql(
			f"SELECT name, description FROM `tabItem` WHERE name IN ({placeholders})",
			_item_codes, as_dict=True
		):
			_item_desc[r_["name"]] = _strip_html(r_.get("description") or "")

	for idx, line in enumerate(lines):
		r  = DATA_START + idx
		rt = line.nl_row_type or "Main Item"
		bg = _ROW_BG.get(rt, "FFFFFF")

		# Row description → item master fallback
		desc_plain = _row_desc.get(line.name, "")
		if not desc_plain and line.item_code:
			desc_plain = _item_desc.get(line.item_code, "")
		if desc_plain:
			n_lines = desc_plain.count('\n') + max(1, len(desc_plain) // 45)
			ws.row_dimensions[r].height = max(14, min(n_lines * 12, 130))
		else:
			ws.row_dimensions[r].height = 16

		for col_l, field, _lbl, _w, is_formula in _COLS:
			cell = ws[f"{col_l}{r}"]

			if is_formula:
				cell.value = _formula(field, r)
			else:
				val = getattr(line, field, None)
				if field == "description":
					val = desc_plain
				cell.value = val if val is not None else ""

			if isinstance(cell.value, str) and cell.value and not cell.value.startswith('='):
				for part in cell.value.split('\n'):
					w = len(part) + 2
					if w > col_max.get(col_l, 0):
						col_max[col_l] = w

			cell.number_format = _num_fmt(field)

			if col_l in _LANDED:   cell.fill = _fill("E8F0FA")
			elif col_l in _SELL:   cell.fill = _fill("E8F5E9")
			elif col_l in _FX:     cell.fill = _fill("FFFFF0")
			else:                  cell.fill = _fill(bg)

			if col_l in _LANDED:
				cell.font = Font(bold=True, size=9, color="0A2A4A")
			elif col_l in _SELL:
				cell.font = Font(bold=True, size=9, color="1A3D1A")
			elif field == "nl_gm_pct":
				gm = flt(getattr(line, "nl_gm_pct", 0))
				fc = "27AE60" if gm >= 30 else "E67E22" if gm >= 20 else "C0392B"
				cell.font = Font(bold=True, size=9, color=fc)
			elif field == "nl_proposed_brand":
				cell.font = Font(bold=True, size=9, color="1A1A3E" if rt != "Accessory" else "555555")
			else:
				cell.font = Font(size=9, color="1A1A3E" if rt != "Accessory" else "555555")

			if field in ("nl_is","nl_price_type","nl_exw_currency","uom","nl_approval_risk"):
				cell.alignment = _align("center")
			elif _num_fmt(field) != "@":
				cell.alignment = _align("right")
			else:
				cell.alignment = _align("left", wrap=field == "description")

			if col_l == "A":
				accent = "1A1A3E" if rt == "Main Item" else "4A7AB5" if rt == "Driver" else "AAAAAA"
				cell.border = _left_accent(accent)
			else:
				cell.border = _thin()

	# Build group index → list of column letters
	_grp_cols = {}
	for gi, (c1, c2, _lbl, _bg, _fg) in enumerate(_GROUPS):
		i1 = column_index_from_string(c1)
		i2 = column_index_from_string(c2)
		_grp_cols[gi] = [get_column_letter(i) for i in range(i1, i2 + 1)]

	for col_l, field, _lbl, width_px, _ in _COLS:
		min_w = max(width_px / 9.0, 5)
		cap   = 44 if field == "description" else 52
		ws.column_dimensions[col_l].width = min(max(col_max.get(col_l, min_w), min_w), cap)

	# Hide columns belonging to collapsed groups
	for gi in hidden_groups:
		for col_l in _grp_cols.get(gi, []):
			ws.column_dimensions[col_l].hidden = True

	N     = DATA_START + len(lines) - 1
	tot_r = N + 1
	ws.row_dimensions[tot_r].height = 18
	tot_fill = _fill("111828")

	for col_l, field, _lbl, _w, _ in _COLS:
		cell = ws[f"{col_l}{tot_r}"]
		cell.fill      = tot_fill
		cell.font      = Font(bold=True, color="FFFFFF", size=9)
		cell.alignment = _align("right")
		cell.border    = Border(
			left=Side(style="thin", color="FFFFFF"),
			right=Side(style="thin", color="FFFFFF"),
			top=Side(style="medium", color="FFFFFF"),
			bottom=Side(style="medium", color="FFFFFF"),
		)
		if field == "nl_is":
			cell.value = "TOTAL"; cell.alignment = _align("left")
		elif field == "nl_exworks_total_aed":
			cell.value = f"=SUM(Q{DATA_START}:Q{N})"; cell.number_format = "#,##0"; cell.font = Font(bold=True, color="7BB8F5", size=9)
		elif field == "nl_landed_total_aed":
			cell.value = f"=SUM(AH{DATA_START}:AH{N})"; cell.number_format = "#,##0"; cell.font = Font(bold=True, color="7BB8F5", size=9)
		elif field == "nl_total_sell_aed":
			cell.value = f"=SUM(AK{DATA_START}:AK{N})"; cell.number_format = "#,##0"; cell.font = Font(bold=True, color="7CF59D", size=9)
		elif field == "qty":
			cell.value = f"=SUM(AM{DATA_START}:AM{N})"; cell.number_format = "#,##0"

	brands = {}
	for line in lines:
		if line.nl_row_type != "Main Item":
			continue
		b = line.nl_supplier_brand or line.nl_proposed_brand or "Unknown"
		if b not in brands:
			brands[b] = {"exw": 0.0, "land": 0.0, "sell": 0.0}
		brands[b]["exw"]  += flt(line.nl_exworks_total_aed)
		brands[b]["land"] += flt(line.nl_landed_total_aed)
		brands[b]["sell"] += flt(line.nl_total_sell_aed)

	if brands:
		bs = tot_r + 2
		ws.merge_cells(f"A{bs}:H{bs}")
		c = ws[f"A{bs}"]
		c.value = "BRAND SUMMARY"; c.fill = _fill("1A1A3E")
		c.font = Font(bold=True, color="FFFFFF", size=11)
		c.alignment = _align("left")
		ws.row_dimensions[bs].height = 22

		hr = bs + 1
		ws.row_dimensions[hr].height = 16
		bs_hdrs = [("A","BRAND"),("B","EXW TOTAL"),("C","% EXW"),
		           ("D","LANDED"),("E","% LND"),("F","PROFIT"),
		           ("G","SELL TOTAL"),("H","% SELL")]
		for ltr, lbl in bs_hdrs:
			c = ws[f"{ltr}{hr}"]
			c.value = lbl; c.fill = _fill("263547")
			c.font = Font(bold=True, color="B8C8E0", size=8)
			c.alignment = _align("center")

		ge = sum(v["exw"]  for v in brands.values())
		gl = sum(v["land"] for v in brands.values())
		gs = sum(v["sell"] for v in brands.values())

		dr = hr + 1
		for b, d in sorted(brands.items()):
			ws.row_dimensions[dr].height = 15
			for ltr, _ in bs_hdrs:
				ws[f"{ltr}{dr}"].fill   = _fill("FFFFFF")
				ws[f"{ltr}{dr}"].border = _thin()
				ws[f"{ltr}{dr}"].font   = Font(size=9)
				ws[f"{ltr}{dr}"].alignment = _align("right")

			profit = d["sell"] - d["land"]
			ws[f"A{dr}"].value = b; ws[f"A{dr}"].font = Font(bold=True, size=9); ws[f"A{dr}"].alignment = _align("left")
			ws[f"B{dr}"].value = d["exw"];   ws[f"B{dr}"].number_format = "#,##0.00"
			ws[f"C{dr}"].value = (d["exw"]/ge*100) if ge else 0; ws[f"C{dr}"].number_format = "0.0"
			ws[f"D{dr}"].value = d["land"];  ws[f"D{dr}"].number_format = "#,##0.00"
			ws[f"E{dr}"].value = (d["land"]/gl*100) if gl else 0; ws[f"E{dr}"].number_format = "0.0"
			ws[f"F{dr}"].value = profit;     ws[f"F{dr}"].number_format = "#,##0.00"; ws[f"F{dr}"].font = Font(bold=True, size=9, color="27AE60")
			ws[f"G{dr}"].value = d["sell"];  ws[f"G{dr}"].number_format = "#,##0"; ws[f"G{dr}"].font = Font(bold=True, size=9)
			ws[f"H{dr}"].value = (d["sell"]/gs*100) if gs else 0; ws[f"H{dr}"].number_format = "0.0"
			dr += 1

		ws.row_dimensions[dr].height = 16
		bdr = Border(top=Side(style="medium", color="263547"),
		             bottom=Side(style="thin", color="D0D0D0"),
		             left=Side(style="thin", color="D0D0D0"),
		             right=Side(style="thin", color="D0D0D0"))
		for ltr, _ in bs_hdrs:
			ws[f"{ltr}{dr}"].fill = _fill("F0F4FF"); ws[f"{ltr}{dr}"].border = bdr
			ws[f"{ltr}{dr}"].font = Font(bold=True, size=9); ws[f"{ltr}{dr}"].alignment = _align("right")
		ws[f"A{dr}"].value = "TOTAL"; ws[f"A{dr}"].alignment = _align("left")
		ws[f"B{dr}"].value = ge;      ws[f"B{dr}"].number_format = "#,##0.00"
		ws[f"C{dr}"].value = 100;     ws[f"C{dr}"].number_format = "0.0"
		ws[f"D{dr}"].value = gl;      ws[f"D{dr}"].number_format = "#,##0.00"
		ws[f"E{dr}"].value = 100;     ws[f"E{dr}"].number_format = "0.0"
		ws[f"F{dr}"].value = gs - gl; ws[f"F{dr}"].number_format = "#,##0.00"; ws[f"F{dr}"].font = Font(bold=True, size=9, color="27AE60")
		ws[f"G{dr}"].value = gs;      ws[f"G{dr}"].number_format = "#,##0"
		ws[f"H{dr}"].value = 100;     ws[f"H{dr}"].number_format = "0.0"

	ws.sheet_view.showGridLines = True
	return wb


@frappe.whitelist()
def download_costing_excel(quotation, hidden_groups=None):
	doc = frappe.get_doc("Quotation", quotation)
	frappe.has_permission("Quotation", doc=doc, throw=True)
	hg = []
	if hidden_groups:
		try:
			hg = [int(x) for x in str(hidden_groups).split(",") if x.strip().isdigit()]
		except Exception:
			hg = []
	wb  = _build_workbook(doc, hidden_groups=hg)
	buf = io.BytesIO()
	wb.save(buf)
	frappe.response["filename"]    = f"NL_Costing_{doc.name}.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"]        = "download"
